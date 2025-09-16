import logging
import os
from openpyxl import Workbook, load_workbook
from config import settings

def export_to_excel(data: list[tuple], sheet_name: str):
    """
    Exports data to a specific sheet in the TimeCollect.xlsx workbook.
    """
    os.makedirs(settings.OUTPUT_DIR, exist_ok=True)

    try:
        wb = load_workbook(settings.EXCEL_OUTPUT_PATH)
    except FileNotFoundError:
        wb = Workbook()

    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])

    ws = wb.create_sheet(sheet_name)
    # Header row matches the TimesheetEntry model
    ws.append([
        "対応",
        "行番号",
        "年",
        "月",
        "日",
        "WeekType",
        "名前",
        "工号",
        "種別",
        "直接/間接",
        "原寸/3D/管理",
        "時間"
    ])

    for row_data in data:
        ws.append(row_data)

    wb.save(settings.EXCEL_OUTPUT_PATH)
    logging.info(f"💾 Data for '{sheet_name}' saved to {settings.EXCEL_OUTPUT_PATH}")