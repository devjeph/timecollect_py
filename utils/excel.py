import os
from os import path
import logging

from openpyxl import Workbook, load_workbook
from dotenv import load_dotenv

load_dotenv()


def export(data, sheet_name):

    save_directory = f"{os.getenv('OUTPUT_DIRECTORY')}"
    file_path = path.join(save_directory, "TimeCollect.xlsx")

    os.makedirs(str(save_directory), exist_ok=True)

    try:
        wb = load_workbook(file_path)
    except FileNotFoundError:
        wb = Workbook()

    if sheet_name in wb.sheetnames:
        wb.remove(wb[sheet_name])

    ws = wb.create_sheet(sheet_name)
    ws.append(
        [
            "対応",
            "行番号",
            "年",
            "月",
            "日",
            "WeekType",
            "名前",
            "工号",
            "種別",
            "直接/間接",
            "原寸/3D/管理",
            "時間",
        ]
    )
    for row in data:
        ws.append(row)

    wb.save(file_path)
    logging.info(f"💾 Filename: TimeCollect.xlsx Saved to {save_directory}.")
